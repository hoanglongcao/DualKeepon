# -*- coding: iso-8859-1 -*-
import time

from openpyxl import load_workbook

########################################
### INITIALISATION TO USE EXCEL FILE ###
########################################

excel_file = load_workbook(filename='Quiz_Multiple_Answers.xlsx')

sheet_quiz = excel_file['Quiz']
sheet_general = excel_file['General']
sheet_keepon = excel_file['Keepon']
sheet_nao = excel_file['Nao']
sheet_nao_voices = excel_file['Nao Voices']

#################################################
### GENERAL FUNCTION (USED ALONG THIS SCRIPT) ###
#################################################

def rowInitial(sheet,types,columnNumber):
    """Return the number of the initial row for the instructions.
    Used for the post questionnaire and for the controller
    Argument:
    - sheet: sheet_questions, sheet_general, etc.
    - types [str]: "Open Question", "Likert type Scale", etc.
    - types [str]: "Right Male & Left Female", "Right Female & Left Male", etc.
	- columnNumber [int]
    Example:
    "Right Male & Left Female" --> i = number of row
    where this string is written in the sheet_general
    """
    i = 0
    a = ''
    while a != types:
        i = i+1
        a = sheet.cell(row = i,column = columnNumber).value
    return i

########################################################
### EXTRACTION OF ALL TEXTS AND PICTURES FOR THE GUI ###
########################################################

def textOutQuiz():
    """Dictionary of all the texts which will be out of the quiz"""
    dictionaryTexts = {}
    option = ""
    i = 0
    while option != "Button":
        option = sheet_general.cell(row = i+1,column = 1).value
        if option != None and option != "Window WELCOME" and option != "Window DATA" and option != "Window RULES" and \
           option != "Window END" and option != "BUTTON" and option != "Window ALERT" and option != "Window QUIZ":
            dictionaryTexts[option] = sheet_general.cell(row = i+1,column = 2).value
        i = i + 1
    return dictionaryTexts

def textQuiz():
    """Dictionary of all the texts which will be in the quiz"""
    dictionaryTexts = {}
    i = rowInitial(sheet_quiz,"Question",2)
    number = 1
    stop = 0
    while stop == 0:
        a = sheet_quiz.cell(row = i+1,column = 2).value
        dictionaryTexts[number] = a
        if type(dictionaryTexts[number]) != unicode:
            del dictionaryTexts[number]
            stop = 1
        number = number + 1
        i = i+1
    return dictionaryTexts

def answerQuiz():
    """List of lists of all the answers for each question of the quiz
    NB:
    FFFFFF00 --> code for the color yellow (left robot)
    else --> for the right robot
    """
    liste = []
    i = 0
    numberAnswer = 0L #to have the variable numberAnswer as a long type
    while type(numberAnswer) == long:
        numberAnswer = sheet_quiz.cell(row = i+4,column = 4).value
        if type(numberAnswer) == long:
            listeAnswer = []
            for j in range(numberAnswer):
                if sheet_quiz.cell(row = i+4,column = 5+j).fill.start_color.index == "FFFFFF00":
                    a = (sheet_quiz.cell(row = i+4,column = 5+j).value,"left")
                else:
                    a = (sheet_quiz.cell(row = i+4,column = 5+j).value,"right")
                listeAnswer.append(a)
            liste.append(listeAnswer)
        i = i+1
    return liste

def pictures():
    """Dictionnary of all pictures for the quiz"""
    dictionaryPictures = {}
    i = 0
    stop = 0
    while stop == 0:
        a = str(sheet_quiz.cell(row = i+4,column = 3).value)
        a = 'Pictures/' + a
        if a == 'Pictures/None':
            stop = 1
        else:
            dictionaryPictures[i] = a
            i = i+1
    return dictionaryPictures

### The four following global variables called the functions above
### They are called in the other scripts to used the values that they store
texts_out = textOutQuiz()
list_pictures = pictures()
messages_quiz = textQuiz()
answers_quiz = answerQuiz()

############################################
### EXTRACTION FOR GENERAL CONFIGURATION ###
############################################

def generalConfiguration(types):
    """
    Dictionnary for the configuration, following the value of "types" [str]
    """
    i = rowInitial(sheet_general,types,4)
    if types == "First number question i-1" or types == "Last number question n":
        i = i - 1
    option = ""
    configurationDictionary = {}
    while option != None:
        option = sheet_general.cell(row = i+1,column = 4).value
        if option != None:
            configurationDictionary[option] = sheet_general.cell(row = i+1,column = 5).value
        i = i + 1
    return configurationDictionary

### The 2 following global variables called the function above
### They are called directly in the function just below
configuration_places = generalConfiguration("Role places")
configuration_robot = generalConfiguration("Robot Choices")

### The 2 following global variables set the number of the first question and
### the total number of questions.
### There are also two global variables just below which have the same names
### but are in comments. They can be used during the development of the code or
### of a quiz, to facilitate the tests
### (indeed you will be able to fix the value of these two variables directly
### in the Excel file and then choose which question you want to test exactly)
question_i = 0
total_number_question = len(messages_quiz)
##question_i = generalConfiguration("First number question i-1")["First number question i-1"]
##total_number_question = generalConfiguration("Last number question n")["Last number question n"]

def definitionConfiguration(dictionary):
    for key,value in dictionary.items():
        if value == 1L:
            return key
    return None

### The 2 following global variables called the function above and
### set the global configuration thanks to the previous global variables 
### configuration_places and configuration_robot
role_definition = definitionConfiguration(configuration_places)
robots_definition = definitionConfiguration(configuration_robot)

### The following global variable stores the date and time of the experiment
liste_time = [time.strftime("%A %d %B %Y"),time.strftime("%H:%M:%S")]

##########################################################################
### EXTRACTION OF BEHAVIORS FOR THE ROBOT CONTROL (VOICES + MOVEMENTS) ###
##########################################################################

def rowInitialNumber(sheet,window):
    """Return the number of the initial row for the instructions
    Arguments:
    - sheet (for example): sheetKeepon = excel_file["Keepon"]
    - window: "Position Initialisation", "Welcome window", "End window"
	or "quiz"
    """
    i = 0
    a = ''
    if window == "quiz":
        while type(a) != long:
            i = i+1
            a = sheet.cell(row = i,column = 1).value
    else:
        while a != window:
            i = i+1
            a = sheet.cell(row = i,column = 1).value
    return i

def numberIterationOther(sheet,rowInit,window):
    """Return the number of lines of window (out quiz) for movements and voices:
    Arguments:
    - sheet (for example): sheetKeepon = excel_file['Keepon']
    - rowInit: number of initial lines,
	calculated by rowInitialNumber(sheet,window)
    - window: "Position Initialisation", "Welcome window" or "End window"
    """
    number = 0
    finish = False
    nextRow = sheet.cell(row = rowInit+1,column = 1).value
    while finish == False:
        column1 = sheet.cell(row = rowInit,column = 1).value
        column2 = sheet.cell(row = rowInit,column = 2).value
        column3 = sheet.cell(row = rowInit,column = 3).value
        if (window == "Position Initialisation" and column1 != window and column1 != None) or \
           (window == "Welcome window" and type(column1) == long) or \
           (window == "End window" and str(column2) == str(column3) == 'None'):
            finish = True
        else:
            number = number + 1
            rowInit = rowInit + 1
    return number

def numberIterationQuiz(sheet,rowInit):
    """Return the number of lines of QUIZ window for movements and voices:
    Arguments:
    - sheet (for example): sheetKeepon = excel_file['Keepon']
    - rowInit: number of initial lines,
	calculated by rowInitialNumber(sheet,window)
    """
    numberQuestion = 0
    numberQuiz = []
    finish = False
    i = rowInit
    while finish == False:
        a = sheet.cell(row = i,column = 1).value
        if a == "End window":
            numberQuiz.append(numberQuestion)
            #we have to repeat this intruction to take into account the last question
            #if there is not this line, the last value of numberQuestion will
            #not be added to numberQuiz
            finish = True
        elif str(a) == "None" or a == 1:
            numberQuestion = numberQuestion + 1
        else:
            numberQuiz.append(numberQuestion)
            numberQuestion = 1
        i = i + 1
    return numberQuiz

def constructionList(i,numberColumn,sheet,rowInit,sheetName):
    """Help of the construction of the list for the MOVEMENTS or VOICES
    of the robot. Used in the function robotData(window,numberColumn,sheetName)
    Return the audio file or the instruction (after transformation if unicode)
    Arguments:
    - i: from the for loop in robotData()
    - numberColumn: 2 (left) or 3 (right) (or 4 (left) or 5 (right) for the end of a question)
    - sheet
    - rowInit: number of initial lines, calculated by rowInitialNumber(sheet,window)
    - sheetName: title of the sheet. For example sheetName = "Keepon Voices"
    """
    a = sheet.cell(row = i+rowInit,column = numberColumn).value
    if type(a) == unicode:
        command = a.encode('utf-8')
    elif type(a) == float or type(a) == long:
        command = a
    else:
        command = "None"
    return command

def robotData(window,numberColumn,sheetName):
    """
    Return the list for the MOVEMENTS or VOICES of the robot following
    if we are in the welcome, quiz or end window.
    Arguments:
    - window: "Position Initialisation", "Welcome window", "End window" or "quiz"
    - numberColumn: 2 (left) or 3 (right) (or 4 (left) or 5 (right) for the end of a question)
    - sheetName: title of the sheet. For example sheetName = "Keepon Voices"
    """
    sheet = excel_file[sheetName]
    listInstruction = []
    listInstructionQuestion = []   
    if window == "quiz":
        rowInit = rowInitialNumber(sheet,window)
        numberIteration = numberIterationQuiz(sheet,rowInit)
        for i in numberIteration:
            for j in range(i):
                instruction = constructionList(j,numberColumn,sheet,rowInit,sheetName)
                if instruction != "None":
                    listInstructionQuestion.append(instruction)
            rowInit = rowInit + i
            listInstruction.append(listInstructionQuestion)
            listInstructionQuestion = []
    else:
        rowInit = rowInitialNumber(sheet,window)
        numberIteration = numberIterationOther(sheet,rowInit,window)
        for i in range(numberIteration):
            instruction = constructionList(i,numberColumn,sheet,rowInit,sheetName)
            if instruction != "None":
                listInstruction.append(instruction)
    return listInstruction

######################################################
### EXTRACTION OF DATA FOR THE ROBOT CONFIGURATION ###
######################################################

def robotConfiguration(sheet,robot,types):
    i = rowInitial(sheet,types,1)
    configurationDictionary = {}
    if robot == "Keepon":
        configurationDictionary["Left Robot USB"] = sheet.cell(row = i,column = 2).value
        configurationDictionary["Right Robot USB"] = sheet.cell(row = i,column = 3).value
    elif robot == "Nao" and types == "USB Port":
        configurationDictionary["Left Robot USB"] = sheet.cell(row = i,column = 2).value
        configurationDictionary["Right Robot USB"] = sheet.cell(row = i,column = 3).value
    elif robot == "Nao" and types == "IP robot":
        configurationDictionary["Left Robot IP"] = sheet.cell(row = i,column = 2).value
        configurationDictionary["Right Robot IP"] = sheet.cell(row = i,column = 3).value
    elif robot == "Nao Voices":
        configurationDictionary["Left Robot Volume"] = sheet.cell(row = i,column = 2).value
        configurationDictionary["Right Robot Volume"] = sheet.cell(row = i,column = 3).value
    return configurationDictionary

### The 2 following global variables called the function above and
### stores the USB configuration for the My Keepon robots
left_keepon_USB = robotConfiguration(sheet_keepon,"Keepon","USB Port")["Left Robot USB"]
right_keepon_USB = robotConfiguration(sheet_keepon,"Keepon","USB Port")["Right Robot USB"]

### The 2 following global variables called the function above and
### stores the volume configuration for the Nao robots
left_nao_volume = int(robotConfiguration(sheet_nao_voices,"Nao Voices","Sound Volume")["Left Robot Volume"])
right_nao_volume = int(robotConfiguration(sheet_nao_voices,"Nao Voices","Sound Volume")["Right Robot Volume"])

#####################################################
### EXTRACTION OF DATA FOR THE POST-QUESTIONNAIRE ###
#####################################################

sheet_questions = excel_file['Post-quest Questions']

def instructions():
    rowInit = rowInitial(sheet_questions,"Instructions in the post-questionnaire",1)
    instruction = sheet_questions.cell(row = rowInit,column = 2).value
    button = sheet_questions.cell(row = 1,column = 5).value
    return {"instruction":instruction,"button":button}

def question(types):
    """
    Return the list of questions.
	Arguments:
    - types = "Open Question", "Multiple Choices Question (multiple answers)",
    "Likert type Scale", etc. following the kind of question
    """
    if types == "Likert type Scale":
        rowInit = rowInitial(sheet_questions,types,1) + 2
    else:
        rowInit = rowInitial(sheet_questions,types,1) + 1
    a = ''
    i = 1
    liste = []
    while a != None:
        a = sheet_questions.cell(row = rowInit+i,column = 2).value
        if a != None:
            liste.append(a)
            i = i + 1
    return liste

def multipleAnswers(types):
    """
    Return the list of the lists of answers of each MCQ.
    """
    rowInit = rowInitial(sheet_questions,types,1) + 1
    a = ''
    i = 1
    totalListe = []
    while a != None:
        a = sheet_questions.cell(row = rowInit+i,column = 3).value
        if a != None:
            j = 0
            liste = []
            b = ''
            while b != None:
                b = sheet_questions.cell(row = rowInit+i,column=3+j).value
                if b != None:
                    liste.append(b)
                    j = j + 1
            totalListe.append(liste)
            i = i + 1
    return totalListe

def numberItem():
    """
    Return the row number at which the likert questions begin
    and return a list of the number of items for each question
    """
    rowInit = rowInitial(sheet_questions,"Number of Items",3)
    a = ''
    i = 1
    liste = []
    while a != None:
        a = sheet_questions.cell(row = rowInit+i,column = 3).value
        if a != None:
            liste.append(a)
            i = i + 1
    return rowInit,liste

def listeItem():
    """
    Return the list of of items for each question
    """
    rowInit,items = numberItem()
    liste = []
    for i in range(len(items)):
        subListe = []
        for j in range(items[i]):
            subListe.append(sheet_questions.cell(row = rowInit+i+1,column = 4+j).value)
        liste.append(subListe)
    return liste

def scaleLikert():
    """
    Return the list of levels of each Likert question
    Return the messages of the configuration of the Likert questionnaire
    """
    rowInit = rowInitial(sheet_questions,"Likert type Scale",1) + 1
    configuration = {"scale":sheet_questions.cell(row = rowInit,column = 3).value,
                     "lowest":sheet_questions.cell(row = rowInit,column = 4).value,
                     "highest":sheet_questions.cell(row = rowInit,column = 5).value}
    return configuration

### The 9 following global variables called the functions above and
### stores the data needed to run the post-questionnaire
dictionary_instructions = instructions()
liste_open = question("Open Question")
liste_multiple = question("Multiple Choices Question (multiple answers)")
liste_single = question("Multiple Choices Question (one answer)")
liste_likert = question("Likert type Scale")
liste_item = listeItem()
liste_multiple_answers = multipleAnswers("Multiple Choices Question (multiple answers)")
liste_single_answer = multipleAnswers("Multiple Choices Question (one answer)")
dictionary_configuration = scaleLikert()
